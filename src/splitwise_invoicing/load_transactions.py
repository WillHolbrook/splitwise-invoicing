import logging
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

from splitwise_invoicing.load_env import environment
from splitwise_invoicing.invoice_extraction import load_amex_credit_transaction_data, \
    load_nationwide_debit_transaction_data, load_hsbc_transaction_data, Currencies

card_name_list, transaction_dates, details, amounts, currencies = [], [], [], [], []


def append_tuple(
        results_tuple: tuple[list[datetime.date], list[str], list[float]],
        card_name: str,
        currency: Currencies,
        card_name_list_p: list[str],
        transaction_dates_p: list[datetime.date],
        details_p: list[str],
        amounts_p: list[float],
        currencies_p: list[str]) -> None:
    """
    Unpacks a result tuple and adds the relevant fields to the 5 lists provided

    Args:
        currencies_p:
        currency:
        results_tuple: The tuple to unpack from load transactions
        card_name:
        card_name_list_p:
        transaction_dates_p:
        details_p:
        amounts_p:

    Returns:

    """
    card_name_list_p += [card_name] * len(results_tuple[0])
    transaction_dates_p += results_tuple[0]
    details_p += results_tuple[1]
    amounts_p += results_tuple[2]
    currencies_p += [currency.value] * len(results_tuple[0])


# Load in data from CSVs into expected format
append_tuple(load_hsbc_transaction_data(Path("./invoices/hsbcCredit/HSBCCredit.csv")),
             "HSBC Credit", Currencies.GBP, card_name_list, transaction_dates, details, amounts, currencies)
append_tuple(load_amex_credit_transaction_data(Path("./invoices/amexCredit/AmexCredit.csv")),
             "Amex Credit", Currencies.GBP, card_name_list, transaction_dates, details, amounts, currencies)
append_tuple(load_nationwide_debit_transaction_data(Path("./invoices/nationwideDebit/NationwideDebit.csv")),
             "Nationwide Debit", Currencies.GBP, card_name_list, transaction_dates, details, amounts, currencies)
append_tuple(load_hsbc_transaction_data(Path("./invoices/hsbcDebit/HSBCDebit.csv")),
             "HSBC Debit", Currencies.GBP, card_name_list, transaction_dates, details, amounts, currencies)

# Specify Excel Column names based off environment vars
transaction_sheet_name = os.getenv("TRANSACTION_SHEET_NAME")
config_sheet_name = os.getenv("CONFIG_SHEET_NAME")
received_date_column_name = os.getenv("RECEIVED_DATE_COLUMN_NAME")
transaction_date_column_name = os.getenv("TRANSACTION_DATE_COLUMN_NAME")
details_column_name = os.getenv("DETAILS_COLUMN_NAME")
amount_column_name = os.getenv("AMOUNT_COLUMN_NAME")
exact_match_column_name = os.getenv("EXACT_MATCH_COLUMN_NAME")
partial_match_column_name = os.getenv("PARTIAL_MATCH_COLUMN_NAME")
partial_match_category_column_name = os.getenv("PARTIAL_MATCH_CATEGORY_COLUMN_NAME")
group_column_name = os.getenv("GROUP_COLUMN_NAME")
add_to_splitwise_column_name = os.getenv("ADD_TO_SPLITWISE_COLUMN_NAME")
card_column_name = os.getenv("CARD_COLUMN_NAME")
adding_default = os.getenv("ADDING_DEFAULT").lower() == "true"

# Specify Excel Formatting
add_to_splitwise_data_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
date_format = NamedStyle("date_format", number_format="d mmmm yyyy")
currency_format = NamedStyle("currency_format", number_format="#,###.00;[Red]-#,###.00;0.00")

logging.info(f"Environment: {environment}")

# Load in template Excel
template_xl: Workbook = load_workbook("./src/splitwise_invoicing/TemplateExcel.xlsx")
transaction_sheet: Worksheet = template_xl[transaction_sheet_name]
config_sheet: Worksheet = template_xl[config_sheet_name]

transaction_column_dict = {}

column_number = 1
for column in transaction_sheet.iter_cols(1, transaction_sheet.max_column):
    transaction_column_dict[column[0].value] = column_number
    column_number += 1


def write_list_to_column(sheet: Worksheet, column_index: int, values: list, excel_format: NamedStyle | None = None) -> None:
    """Writes the content of a list to a given column index in the specified worksheet applying a style if given"""
    row_num = 2
    for value in values:
        current_cell = sheet.cell(row=row_num, column=column_index)
        current_cell.value = value
        if excel_format:
            current_cell.style = excel_format

        row_num += 1


num_transactions = len(transaction_dates)
default_group = config_sheet.cell(2, 1).value

# Writing data into Excel sheet
write_list_to_column(transaction_sheet, transaction_column_dict[card_column_name], card_name_list)
write_list_to_column(transaction_sheet, transaction_column_dict[group_column_name], [default_group] * num_transactions)
write_list_to_column(transaction_sheet, transaction_column_dict[add_to_splitwise_column_name],
                     [adding_default] * num_transactions)
write_list_to_column(transaction_sheet, transaction_column_dict[transaction_date_column_name], transaction_dates,
                     date_format)
write_list_to_column(transaction_sheet, transaction_column_dict[details_column_name], details)
write_list_to_column(transaction_sheet, transaction_column_dict[amount_column_name], amounts, currency_format)
write_list_to_column(transaction_sheet, transaction_column_dict["Currency"], currencies)

# Adding data validation onto specific column
ascii_uppercase_offset = 64
add_to_splitwise_column_letter = chr(transaction_column_dict[add_to_splitwise_column_name] + ascii_uppercase_offset)
add_to_splitwise_data_validation.add(f'{add_to_splitwise_column_letter}2:{add_to_splitwise_column_letter}1048576')
transaction_sheet.add_data_validation(add_to_splitwise_data_validation)

# Output the final file
template_xl.save(f"SplitwiseInvoicing-{str(datetime.now()).replace(':', '.')}.xlsx")
