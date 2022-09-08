import logging
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.worksheet import Worksheet

from splitwise_invoicing.load_env import environment
from splitwise_invoicing.invoice_extraction import load_hsbc_transaction_data


hsbc_path = Path(os.getenv("HSBC_CREDIT_INVOICE_FILEPATH"))
received_dates, transaction_dates, details, amounts = load_hsbc_transaction_data(hsbc_path)

template_xl_path = Path("./TemplateExcel.xlsx")
transaction_sheet_name = "Transactions"
config_sheet_name = "Config"
received_date_column_name = "Received Date"
transaction_date_column_name = "Transaction Date"
details_column_name = "Details"
amount_column_name = "Amount"
exact_match_column_name = "Exact Match On Splitwise?"
partial_match_column_name = "Partial Match on Splitwise?"
partial_match_category_column_name = "Partial Match Category"

date_format = NamedStyle("date_format", number_format="d mmmm yyyy")
currency_format = NamedStyle("currency_format", number_format="£#,###.00;[Red]£-#,###.00;0.00")

logging.info(f"Environment: {environment}")

template_xl: Workbook = load_workbook(template_xl_path)
transaction_sheet: Worksheet = template_xl[transaction_sheet_name]

transaction_column_dict = {}

column_number = 1
for column in transaction_sheet.iter_cols(1, transaction_sheet.max_column):
    transaction_column_dict[column[0].value] = column_number
    column_number += 1


def write_list_to_column(sheet: Worksheet, column_index: int, values: list, format: NamedStyle | None = None) -> None:
    row_num = 2
    for value in values:
        current_cell = sheet.cell(row=row_num, column=column_index)
        current_cell.value = value
        if format:
            current_cell.style = format

        row_num += 1


write_list_to_column(transaction_sheet, transaction_column_dict[transaction_date_column_name], transaction_dates,
                     date_format)
write_list_to_column(transaction_sheet, transaction_column_dict[received_date_column_name], received_dates, date_format)
write_list_to_column(transaction_sheet, transaction_column_dict[details_column_name], details)
write_list_to_column(transaction_sheet, transaction_column_dict[amount_column_name], amounts, currency_format)

template_xl.save(f"SplitwiseInvoicing-{str(datetime.now()).replace(':', '.')}.xlsx")
